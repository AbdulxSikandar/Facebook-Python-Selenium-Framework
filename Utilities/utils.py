import inspect
import logging
from openpyxl import Workbook, load_workbook


class Utils():

    def getLogger(self):
            loggerName = inspect.stack()[1][3]
            logger = logging.getLogger(loggerName)
            fileHandler = logging.FileHandler('logfile.log')
            formatter = logging.Formatter("%(asctime)s :%(levelname)s : %(name)s :%(message)s")
            fileHandler.setFormatter(formatter)
            logger.addHandler(fileHandler)  # filehandler object
            logger.setLevel(logging.DEBUG)
            return logger

    def read_data_from_excel(self):
            datalist = []
            wb = load_workbook(filename= "K:\\Data Science\Python\\Python Practice\\Project1\\testdata\\testdata.xlsx")
            sh = wb.active
            row_ct = sh.max_row
            col_ct = sh.max_column
            for i in range(1, row_ct + 1):
                row = []
                for j in range(2, col_ct + 1):
                    row.append(sh.cell(row=i, column=j).value)
                    datalist.append(row)
            return datalist
